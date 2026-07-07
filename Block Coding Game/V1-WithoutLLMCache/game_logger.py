"""
Excel game logging — two-player sessions with ArUco ID tracking.

Sheets:
  Sessions      — one row per game (both players, map, outcome, duration)
  Attempts      — one row per RFID submission at a checkpoint
  PlayerSummary — one row per game, rolling up accuracy and score
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

LOG_PATH = Path(__file__).parent / "game_log.xlsx"

SESSIONS_HEADER = [
    "SessionID",
    "Player1 Name", "Player1 ArUco", "Player1 Age", "Player1 Plays",
    "Player2 Name", "Player2 ArUco", "Player2 Age", "Player2 Plays",
    "Date", "Map", "Start Time", "End Time", "Duration (s)", "Outcome",
]

ATTEMPTS_HEADER = [
    "SessionID", "Player1 Name", "Player2 Name",
    "Date", "Map", "Checkpoint", "Attempt #",
    "Scanned RFIDs", "Expected RFIDs", "Result", "Correct?",
    "Start Time", "End Time", "Duration (s)",
]

SUMMARY_HEADER = [
    "SessionID",
    "Player1 Name", "Player1 ArUco",
    "Player2 Name", "Player2 ArUco",
    "Date", "Map", "Outcome",
    "Total Attempts", "Correct Attempts", "Incorrect Attempts",
    "Accuracy (%)", "Game Duration (s)",
]


def _ensure_workbook() -> Workbook:
    if LOG_PATH.exists():
        wb = load_workbook(LOG_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Sessions" not in wb.sheetnames:
        wb.create_sheet("Sessions").append(SESSIONS_HEADER)
    if "Attempts" not in wb.sheetnames:
        wb.create_sheet("Attempts").append(ATTEMPTS_HEADER)
    if "PlayerSummary" not in wb.sheetnames:
        wb.create_sheet("PlayerSummary").append(SUMMARY_HEADER)

    return wb


class GameLogger:
    """One instance per game. Call start(), log_attempt()*, end()."""

    def __init__(self, players: list[dict], map_name: str):
        """
        players: list of 2 dicts from id_scanner.wait_for_players()
                 each has keys: aruco_id, name, age, plays
        """
        self.players  = players
        self.map_name = map_name
        self.session_id: str | None = None
        self._session_start: datetime | None = None
        self._checkpoint_start: datetime | None = None
        self._attempt_rows: list[dict] = []

    def _p(self, idx: int) -> dict:
        return self.players[idx] if idx < len(self.players) else {}

    # ── session-level ──────────────────────────────────────────────────────

    def start(self):
        self._session_start = datetime.now()
        p1 = self._p(0)
        p2 = self._p(1)
        self.session_id = (
            f"{self._session_start.strftime('%Y%m%dT%H%M%S')}"
            f"_{p1.get('aruco_id', 'X')}_{p2.get('aruco_id', 'X')}"
        )
        print(f"  Session ID: {self.session_id}")

    def end(self, outcome: str):
        end_time = datetime.now()
        start    = self._session_start or end_time
        duration = round((end_time - start).total_seconds(), 1)
        p1, p2   = self._p(0), self._p(1)

        wb = _ensure_workbook()

        wb["Sessions"].append([
            self.session_id,
            p1.get("name", ""),  p1.get("aruco_id", ""), p1.get("age", ""), p1.get("plays", 0),
            p2.get("name", ""),  p2.get("aruco_id", ""), p2.get("age", ""), p2.get("plays", 0),
            start.strftime("%Y-%m-%d"),
            self.map_name,
            start.strftime("%H:%M:%S"),
            end_time.strftime("%H:%M:%S"),
            duration,
            outcome,
        ])

        total    = len(self._attempt_rows)
        correct  = sum(1 for a in self._attempt_rows if a["correct"])
        accuracy = round(100 * correct / total, 1) if total else 0.0

        wb["PlayerSummary"].append([
            self.session_id,
            p1.get("name", ""), p1.get("aruco_id", ""),
            p2.get("name", ""), p2.get("aruco_id", ""),
            start.strftime("%Y-%m-%d"),
            self.map_name,
            outcome,
            total, correct, total - correct, accuracy, duration,
        ])

        wb.save(LOG_PATH)

    # ── per-checkpoint-attempt level ──────────────────────────────────────

    def begin_checkpoint_attempt(self):
        self._checkpoint_start = datetime.now()

    def log_attempt(self, checkpoint_label: str, attempt_num: int,
                    scanned: list[int], expected: list[int], result: str):
        end_time = datetime.now()
        start    = self._checkpoint_start or end_time
        duration = round((end_time - start).total_seconds(), 1)
        correct  = (result == "CORRECT")

        self._attempt_rows.append({"correct": correct})

        p1, p2 = self._p(0), self._p(1)
        wb = _ensure_workbook()
        wb["Attempts"].append([
            self.session_id,
            p1.get("name", ""), p2.get("name", ""),
            start.strftime("%Y-%m-%d"),
            self.map_name,
            checkpoint_label,
            attempt_num,
            str(scanned),
            str(expected),
            result,
            "Yes" if correct else "No",
            start.strftime("%H:%M:%S"),
            end_time.strftime("%H:%M:%S"),
            duration,
        ])
        wb.save(LOG_PATH)
